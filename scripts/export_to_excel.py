#!/usr/bin/env python3
"""
Export all source strings + target translations of a Crowdin project to Excel.

Usage:
    python3 export_to_excel.py --project 744513 --lang en --out /root/xagone_en.xlsx

For each source string we join the target translation by stringId.
Plural source strings (dict instead of str) are JSON-serialized.
"""
import argparse
import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent))
from _common import get, paginate, XAG_PROJECTS


def fetch_labels(project_id):
    labels = {}
    for item in paginate(f"/projects/{project_id}/labels"):
        d = item["data"]
        labels[d["id"]] = d["title"]
    return labels


def fetch_source_strings(project_id):
    """Returns {stringId: {identifier, text, context, type, labels, labelNames}}"""
    out = {}
    for item in paginate(f"/projects/{project_id}/strings"):
        d = item["data"]
        text = d.get("text", "")
        # Plurals: source may be {"one": "...", "other": "..."}
        if isinstance(text, dict):
            text = json.dumps(text, ensure_ascii=False)
        out[d["id"]] = {
            "identifier": d.get("identifier", ""),
            "text": text,
            "context": d.get("context", "") or "",
            "type": d.get("type", "text"),
            "labelIds": d.get("labelIds", []),
            "labelNames": [],  # filled later
        }
    return out


def fetch_translations(project_id, lang):
    """Returns {stringId: translation_text} (plurals as JSON string)."""
    out = {}
    for item in paginate(f"/projects/{project_id}/languages/{lang}/translations"):
        d = item["data"]
        sid = d.get("stringId")
        if sid is None:
            continue
        if "text" in d:
            out[sid] = d["text"]
        elif "plural" in d:
            out[sid] = json.dumps(d["plural"], ensure_ascii=False)
        else:
            out[sid] = ""
    return out


def fetch_project(project_id):
    data = get(f"/projects/{project_id}")
    if not data:
        sys.exit(f"Project {project_id} not found or no access.")
    return data["data"]


def build_excel(project_id, lang, out_path):
    project = fetch_project(project_id)
    print(f"Project: {project['name']} ({project['identifier']})")
    print(f"  Source: {project['sourceLanguageId']}, Targets: {len(project['targetLanguageIds'])}")

    print("\n[1/4] Fetching labels...")
    labels = fetch_labels(project_id)
    print(f"  {len(labels)} labels")

    print("\n[2/4] Fetching source strings...")
    sources = fetch_source_strings(project_id)
    print(f"  {len(sources)} source strings")
    # Backfill label names
    for s in sources.values():
        s["labelNames"] = [labels.get(lid, f"#{lid}") for lid in s["labelIds"]]

    print(f"\n[3/4] Fetching {lang} translations...")
    translations = fetch_translations(project_id, lang)
    print(f"  {len(translations)} translations")

    print("\n[4/4] Building Excel...")
    wb = openpyxl.Workbook()

    # Main sheet
    ws = wb.active
    ws.title = f"{project['identifier']}_{lang}"[:31]
    headers = [
        "ID", "Identifier", "Source (zh-CN)", f"Translation ({lang})",
        "Context", "Type", "Labels",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E5A87")
        c.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    matched = 0
    for sid in sorted(sources.keys()):
        src = sources[sid]
        tgt = translations.get(sid, "")
        if tgt:
            matched += 1
        ws.cell(row=row, column=1, value=sid)
        ws.cell(row=row, column=2, value=src["identifier"])
        ws.cell(row=row, column=3, value=src["text"])
        ws.cell(row=row, column=4, value=tgt)
        ws.cell(row=row, column=5, value=src["context"])
        ws.cell(row=row, column=6, value=src["type"])
        ws.cell(row=row, column=7, value=", ".join(src["labelNames"]))
        for c in range(2, 8):
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Column widths
    widths = [8, 40, 60, 60, 30, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary = [
        ("Project", project["name"]),
        ("Project ID", project_id),
        ("Identifier", project["identifier"]),
        ("Source Language", project["sourceLanguageId"]),
        ("Target Language", lang),
        ("Total Source Strings", len(sources)),
        ("Total Translations", len(translations)),
        ("Matched (stringId)", matched),
        ("Unmatched", len(sources) - matched),
        ("Coverage", f"{matched/len(sources)*100:.1f}%" if sources else "0%"),
        ("All Labels", ", ".join(sorted(set(labels.values())))),
    ]
    for r, (k, v) in enumerate(summary, 1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 80

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\nSaved: {out_path} ({out_path.stat().st_size:,} bytes)")
    print(f"  Rows: {row-2}, Matched: {matched}, Unmatched: {len(sources)-matched}")


def main():
    ap = argparse.ArgumentParser(description="Export Crowdin project to Excel")
    ap.add_argument("--project", type=int, required=True,
                    help="Crowdin project ID (e.g. 744513 for xagone)")
    ap.add_argument("--lang", default="en", help="Target language code (default: en)")
    ap.add_argument("--out", required=True, help="Output xlsx path")
    args = ap.parse_args()

    if args.project not in XAG_PROJECTS and not get(f"/projects/{args.project}"):
        sys.exit(f"Project {args.project} not found or inaccessible.")
    build_excel(args.project, args.lang, args.out)


if __name__ == "__main__":
    main()
